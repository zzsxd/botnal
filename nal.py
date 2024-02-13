import asyncio

from aiogram.fsm.state import StatesGroup, State
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from aiogram.types import InputFile, FSInputFile
from aiogram import Bot, Dispatcher, types, F
from aiogram.filters.command import Command
# Включаем логирование, чтобы не пропустить важные сообщения
from aiogram.fsm.context import FSMContext

import logging


logging.basicConfig(level=logging.INFO)
# Объект бота
API_KEY = '4HKE5XT3J0C05PVA'
#6612735530:AAHLsuxpu3J_r1T6_xnO6KfcFTfwoIXCPLg - ye;
#5526766002:AAFGazXTruzsDGJAOfFHhjVUyb8jLij5AlY
bot = Bot(token="5295859051:AAE7QjHay-eFwwfgRkdCrUlky7H0PFKacTI")
# Диспетчер
dp = Dispatcher()




class TextPos(StatesGroup):
    name = State()


async def start(message: types.Message, state: FSMContext):
    kb = [
        [
            types.KeyboardButton(text="Получить накладную"),
        ],
    ]
    keyboard = types.ReplyKeyboardMarkup(
        keyboard=kb,
        resize_keyboard=True,
    )
    await message.answer("Добро пожаловать 👋", reply_markup=keyboard)


@dp.message(F.text == 'Получить накладную')
async def get_text(message: types.Message, state: FSMContext):
    await state.set_state(TextPos.name)
    await message.answer("Введите все нужные данные по порядку:\n"
                         "Дата отправление\nНомер накладной\nНазвание товара\nВес\nОбъем\nКол во\nСтоимость\nУпаковка\nСтраховка\nПрр москва")


@dp.message(TextPos.name)
async def name(message: types.Message, state: FSMContext):
    await state.clear()
    wb = load_workbook('./test.xlsx')

    ws = wb.active
    answer_list = message.text.split('\n')
    moscow = round(float(answer_list[3]) * float(answer_list[6].replace('$', '')), 2)
    ws['B7'] = answer_list[0]
    ws['F7'] = answer_list[1]
    ws['B11'] = answer_list[2]
    ws['D13'] = answer_list[3]
    ws['D15'] = answer_list[4]
    ws['D17'] = answer_list[5]
    ws['D19'] = answer_list[6]
    ws['D21'] = answer_list[7]
    ws['D23'] = answer_list[8]
    ws['D26'] = str(moscow) + '$'
    ws['D28'] = answer_list[9]
    ws['A29'] = f'Итого к оплате {round(float(moscow) + float(answer_list[7].replace("$", "")) + float(answer_list[8].replace("$", "")) + float(answer_list[9].replace("$", "")), 2)}$'

    wb.save(f'./{answer_list[1]}.xlsx')
    await message.answer_document(FSInputFile(f'./{answer_list[1]}.xlsx'))


async def main():
    dp.message.register(start, Command("start"))
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())
